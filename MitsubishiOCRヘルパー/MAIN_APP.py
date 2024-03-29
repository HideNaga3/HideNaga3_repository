import os
from pathlib import Path
import sys
from PyQt5 import QtGui, QtWidgets
from PyQt5.QtCore import Qt, QCoreApplication, QSharedMemory
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog, QDialog
from subprocess import Popen
import win32com.client as win32
from pypdf import PdfReader
import openpyxl
import datetime as dt
import time
import json
from main_app_ui import Ui_MainWindow
from progress_bar_ui import Ui_ProgressBarDialog
from progress_msg_ui import Ui_ProgressMsgDialog
import pprint

class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setAcceptDrops(True)
        self.init_config() # myfunc

        # タブウィジェット
        self.tabWidget: QtWidgets.QTabWidget

        # PDF情報取得タブ内
        # list
        self.page_count_listWidget: QtWidgets.QListWidget # list
        self.pcnt_list_delete_pushButton: QtWidgets.QPushButton
        self.pcnt_listwidget_clear_pushButton: QtWidgets.QPushButton
        self.file_dialog_pcnt_pushButton: QtWidgets.QPushButton
        self.page_count_pushButton: QtWidgets.QPushButton # 実行ボタン
        self.page_count_from_dir_pushButton: QtWidgets.QPushButton # フォルダから実行ボタン
        self.dir_dialog_pcnt_pushButton: QtWidgets.QPushButton # フォルダ選択ボタン
        self.pcnt_label_clear_pushButton: QtWidgets.QPushButton # ラベルクリアボタン

        self.copy_dir_exe_pushButton: QtWidgets.QPushButton # 実行ボタン
        self.recursive_checkBox: QtWidgets.QCheckBox # 再帰的にフォルダを検索するかどうか
        self.page_count_label: QtWidgets.QLabel # フォルダパスラベル

        self.radioButton_pcnt: QtWidgets.QRadioButton
        self.radioButton_psize: QtWidgets.QRadioButton

        # listwidgetを複数選択可能にする
        self.page_count_listWidget.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        # labelを選択可能にする
        self.page_count_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        # イベントドリブン
        self.page_count_pushButton.clicked.connect(lambda: self.count_page_from_list())
        self.dir_dialog_pcnt_pushButton.clicked.connect(lambda: self.set_path(self.page_count_label))
        self.page_count_from_dir_pushButton.clicked.connect(lambda: self.main_process()) # 主処理
        self.pcnt_list_delete_pushButton.clicked.connect(lambda: self.delete_one_item_from_listwidget())

        self.pcnt_listwidget_clear_pushButton.clicked.connect(lambda: self.listwidget_clear(self.page_count_listWidget))
        self.pcnt_label_clear_pushButton.clicked.connect(lambda: self.label_clear(self.page_count_label))

        self.file_dialog_pcnt_pushButton.clicked.connect(lambda: self.set_file_path(self.page_count_listWidget))
        self.recursive_checkBox.stateChanged.connect(lambda: self.change_value_config(self.recursive_checkBox))
        self.radioButton_pcnt.clicked.connect(lambda: self.change_value_config(self.radioButton_pcnt))
        self.radioButton_psize.clicked.connect(lambda: self.change_value_config(self.radioButton_psize))

    ############################## //// 共通設定 ############################

    def change_value_config(self, target_obj):
        config = self.get_config()
        if target_obj == self.recursive_checkBox: # 再帰的にフォルダを検索するかどうか
            if self.recursive_checkBox.isChecked():
                config['is_recursive'] = True
            else:
                config['is_recursive'] = False
        elif target_obj == self.radioButton_pcnt: # ページ数取得モード
            if self.radioButton_pcnt.isChecked():
                config['get_pagecnt'] = True
        elif target_obj == self.radioButton_psize: # ページサイズ取得モード
            if self.radioButton_psize.isChecked():
                config['get_pagecnt'] = False
        self.set_config(config)

    # ラベルをクリア
    def label_clear(self, target_label: QtWidgets.QLabel) -> None:
        if QMessageBox.question(self, 'Confirm', 'フォルダパスをクリアしますか？', QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return
        target_label.setText('')

    # listをすべてクリア
    def listwidget_clear(self, target_list: QtWidgets.QListWidget) -> None:
        if QMessageBox.question(self, 'Confirm', 'リストをクリアしますか？', QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return
        target_list.clear()

    # listの選択項目を削除
    def delete_one_item_from_listwidget(self):
        for item in self.page_count_listWidget.selectedItems():
            self.page_count_listWidget.takeItem(self.page_count_listWidget.row(item))
            if self.page_count_listWidget.count() != 0:
                self.page_count_listWidget.setCurrentRow(0)

    # プログレスバーを表示
    def show_progress_bar(self, max_value):
        self.progress_bar_instance = ProgressBarDialog(max_value) # インスタンス生成
        self.progress_bar_instance.show() # 表示
        QCoreApplication.processEvents() # プロセス更新

    # 現在時刻文字列を取得する
    def get_datetime_str(self):
        return dt.datetime.now().strftime('%Y%m%d_%H%M_%S')

    # jsonファイルを読み込む
    def get_config(self):
        config_path = './data/config.json'
        with open(config_path, 'r') as f:
            config = json.load(f)
        return config

    # jsonファイルに保存
    def set_config(self, config):
        config_path = './data/config.json'
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=4)

    # フォルダパスを設定する
    def set_path(self, target_label: QtWidgets.QLabel) -> None:
        p = QFileDialog.getExistingDirectory(self, 'フォルダを選択', './')
        if p == '':
            return
        p = os.path.normpath(p)
        target_label.setText(p)
        QMessageBox.information(self, 'Info', 'フォルダパスを設定しました')

    # ファイルパスを設定する
    def set_file_path(self, taget_list: QtWidgets.QListWidget) -> None:
        filepaths, _ = QFileDialog.getOpenFileNames(self, 'ファイルを選択', './', 'PDFファイル(*.pdf)')
        if len(filepaths) == 0:
            return
        for i, filepath in enumerate(filepaths):
            filename = os.path.basename(filepath)

            taget_list.addItem(f'{str(i)}\t{filename}\t{filepath}')

    # 初期設定
    def init_config(self):
        config = self.get_config()
        self.page_count_label.setText(config['path']['dir_path_of_pcnt'])

        self.tabWidget.setCurrentIndex(config['init_tab_index'])
        if config['is_recursive'] == True:
            self.recursive_checkBox.setChecked(True)
        else:
            self.recursive_checkBox.setChecked(False)
        if config['get_pagecnt'] == True:
            self.radioButton_pcnt.setChecked(True)
        else:
            self.radioButton_psize.setChecked(True)

    # 終了処理
    def closeEvent(self, event):
        config = self.get_config()
        # フォルダパスを保存
        config['path']['dir_path_of_pcnt'] = self.page_count_label.text()
        # 初期タブを保存
        config['init_tab_index'] = self.tabWidget.currentIndex()
        self.set_config(config)
        event.accept()

    ############################## 共通設定 //// ############################
    ############################### //// PDFのページ数を取得 ##################################
    # フォルダからPDFファイルのページ数を取得
    def main_process(self):
        if self.radioButton_pcnt.isChecked():
            self.count_page_from_dir()
        elif self.radioButton_psize.isChecked():
            self.get_page_size_driver()

    def count_page_from_dir(self): # Driver
        dir_path = self.page_count_label.text() # フォルダパスはラベルに表示されているものを取得
        if dir_path == '':
            QMessageBox.warning(self, 'Warning', 'フォルダが選択されていません')
            return
        elif os.path.exists(dir_path) == False:
            QMessageBox.warning(self, 'Warning', 'フォルダが存在しません')
            return
        if QMessageBox.question(self, 'Confirm', 'PDFのページ数を取得しますか？', QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return
        # 処理ファイル数を取得
        file_cnt = 0
        for i, plib in enumerate(Path(dir_path).glob('**/*.pdf')):
            file_cnt += 1
        # プログレスバーを表示
        self.show_progress_bar(file_cnt)
        self.buf_li = []
        if self.recursive_checkBox.isChecked(): # 再帰的にフォルダを検索するかどうか
            path_collecton = Path(dir_path).glob('**/*.pdf')
        else:
            path_collecton = Path(dir_path).glob('*.pdf')

        for i, plib in enumerate(path_collecton):
            buf_path = str(plib) # ファイルパス
            parent_dir = os.path.dirname(buf_path) # 親ディレクトリパス
            grandparent_dir = os.path.dirname(parent_dir) # 親ディレクトリの親ディレクトリパス
            parent_name = os.path.basename(parent_dir) # 親ディレクトリ名
            grandparent_name = os.path.basename(grandparent_dir) # 親ディレクトリの親ディレクトリ名
            buf_index = i + 1
            ############## 主処理 ##############

            buf_page_cnt = self.count_pdf_page(buf_path) # myfunc PDFのページ数を取得

            ############## 主処理 ##############
            buf_size = os.path.getsize(buf_path) # ファイルサイズ
            buf_name = plib.name # ファイル名
            self.buf_li.append((buf_index, buf_page_cnt, buf_size, grandparent_name, parent_name, buf_name, buf_path))
            # プログレスバーを更新
            self.progress_bar_instance.update_progress_bar(i + 1)
        self.progress_bar_instance.close()
        self.write_excel(self.buf_li) # myfunc

    # リストからPDFファイルのページ数を取得
    def count_page_from_list(self):
        tgt_list_widget = self.page_count_listWidget
        if tgt_list_widget.count() == 0:
            QMessageBox.warning(self, 'Warning', 'PDFファイルが選択されていません')
            return
        if QMessageBox.question(self, 'Confirm', 'PDFのページ数を取得しますか？', QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return
        # プログレスバーを表示
        max_cnt = tgt_list_widget.count()
        progress_bar_instance = ProgressBarDialog(max_cnt)
        progress_bar_instance.show()

        result_list = []
        for i in range(tgt_list_widget.count()):
            buf_path = tgt_list_widget.item(i).text().split('\t')[2]
            buf_name = tgt_list_widget.item(i).text().split('\t')[1]
            buf_index = tgt_list_widget.item(i).text().split('\t')[0]
            parent_dir = os.path.dirname(buf_path)
            grandparent_dir = os.path.dirname(parent_dir)
            parent_name = os.path.basename(parent_dir)
            grandparent_name = os.path.basename(grandparent_dir)

            if buf_path == '' or buf_path.endswith('.pdf') == False:
                continue
            # ファイルサイズを取得
            buf_size = os.path.getsize(buf_path)
            # PDFファイルのページ数を取得
            buf_page_cnt = self.count_pdf_page(buf_path) # myfunc

            # リストに追加
            result_list.append((int(buf_index), buf_page_cnt, int(buf_size), grandparent_name, parent_name, buf_name, buf_path))
            progress_bar_instance.update_progress_bar(i + 1)

        progress_bar_instance.close()
        self.write_excel(result_list) # myfunc

    # エクセルに書き出し
    def write_excel(self, result_list: list) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        columns_name = ['index', 'page_count', 'Byte', 'grandparent_dir', 'parent_dir', 'file_name', 'file_path']
        for i in range(len(columns_name)):
            ws.cell(row=1, column=i + 1).value = columns_name[i]
        for row in range(len(result_list)):
            for col in range(len(result_list[0])):
                ws.cell(row=row + 2, column=col + 1).value = result_list[row][col]
        dt_str = dt.datetime.now().strftime('%Y%m%d_%H%M_%S')
        excel_name = f'./output/_page_count_{dt_str}.xlsx'

        # 列の自動調整
        self.auto_fit_column(ws) # myfunc

        # outputフォルダがなければ作成
        if os.path.exists('./output') == False:
            os.makedirs('./output')
        excel_full_path = os.path.join(os.getcwd(), excel_name)
        if os.path.exists(excel_name) == False:
            wb.save(excel_name)
            if QMessageBox.question(self, 'Confirm',
                                    f'{excel_full_path}\nを作成しました\n\n'
                                    'エクセルファイルを開きますか？',
                                    QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                # 保存したエクセルファイルを開く
                excel_app = win32.gencache.EnsureDispatch('Excel.Application')
                wb = excel_app.Workbooks.Open(excel_full_path)
                excel_app.Visible = True
            else:
                return
        else:
            QMessageBox.warning(self, 'Warning', 'Exctlファイルを書き出せませんでした')
            return

    # 列の自動調整
    def auto_fit_column(self, ws):
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_lenght = len(str(cell.value))
            adjust_width = int((max_lenght + 5) * 1.2)
            ws.column_dimensions[column_letter].width = adjust_width

    # PDFファイルのページ数を取得
    def count_pdf_page(self, file_path: str) -> int:
        if file_path == '' or file_path.endswith('.pdf') == False:
            return 0
        try:
            reader = PdfReader(file_path)
            return len(reader.pages)
        except Exception as e:
            print('Exception:', e)
            return -1
    ############ //// ページサイズ取得 ############
    # 主処理
    def get_page_size_driver(self):
        if QMessageBox.question(self, 'Confirm', 'PDFのページサイズを取得しますか？',
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return
        parent_pstr = self.page_count_label.text()
        if parent_pstr == '':
            QMessageBox.warning(self, 'Warning', 'フォルダが選択されていません')
            return
        if os.path.exists(parent_pstr) == False:
            QMessageBox.warning(self, 'Warning', 'フォルダが存在しません')
            return
        if self.recursive_checkBox.isChecked():
            plist = list(Path(parent_pstr).glob('**/*.pdf'))
        else:
            plist = list(Path(parent_pstr).iterdir())
        self.data = []
        self.show_progress_bar(len(plist)) # myfunc # プログレスバーを表示
        for i, p in enumerate(plist):
            self.progress_bar_instance.update_progress_bar(i + 1) # myfunc # プログレスバーを更新
            if p.is_dir():
                continue
            pstr = str(p)
            if os.path.splitext(pstr)[1] == '.pdf':
                self.get_page_size_helper(pstr) # myfunc
        self.data.pop(0)
        self.progress_bar_instance.close() # プログレスバーを閉じる
        self.write_size_to_xl() # myfunc # エクセルに書き出す

    # ページサイズ取得ヘルパー関数
    def get_page_size_helper(self, pstr: str):
        with open(pstr, 'rb') as f:
            filename = os.path.basename(pstr)
            fullfilepath = os.path.abspath(pstr)
            reader = PdfReader(f)
            pages = reader.pages
            page_count = len(pages)
            self.data.append(['','','','',''])
            self.data.append([fullfilepath, '', '', '', ''])
            self.data.append(['総ページ数 ->', page_count, '', '', ''])
            self.data.append(['filename', 'page', 'w x h', 'width(mm)', 'height(mm)'])
            for i in range(page_count):
                page = pages[i]
                rotate = page.get('/Rotate', 0)
                rotate = ((rotate % 360) + 360) % 360 # 角度正規化
                size = page.mediabox
                width = size.width
                height = size.height
                width_mm = self.points_to_mm(width) # myfunc
                height_mm = self.points_to_mm(height) # myfunc
                if rotate in (90, 270):
                    width_mm, height_mm = height_mm, width_mm
                self.data.append([filename, i + 1, f'{width_mm} x {height_mm} mm', width_mm, rotate])

    def points_to_mm(self, points):
        return int(points * (25.4 / 72))

    # excelで書き出す
    def write_size_to_xl(self):
        dstr = self.get_datetime_str()
        xl_pstr_rel = f'./output/_page_size_{dstr}.xlsx'
        self.xl_pstr = os.path.abspath(xl_pstr_rel)
        if os.path.exists(self.xl_pstr) == False:
            openpyxl.Workbook().save(self.xl_pstr)
        wb = openpyxl.load_workbook(self.xl_pstr)
        ws = wb.active
        # シートをクリア
        ws.delete_rows(1, ws.max_row)
        # for j, column in enumerate(columns):
        #     ws.cell(row = 1, column = j + 1, value = column)
        for i, row in enumerate(self.data):
            for j, cell in enumerate(row):
                ws.cell(row = i + 1, column = j + 1, value = cell)
        column_width_dict = {'A': 58.45, 'B': 6.45, 'C': 15.55}
        for col, width in column_width_dict.items():
            ws.column_dimensions[col].width = width
        try:
            wb.save(self.xl_pstr)
        except:
            print('write error')
            wb.close()
            return
        wb.close()
        if QMessageBox.question(self, 'Confirm',
                                f'{self.xl_pstr}\nを作成しました\n\nエクセルファイルを開きますか？',
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            # 保存したエクセルファイルを開く
            excel_app = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel_app.Workbooks.Open(self.xl_pstr)
            excel_app.Visible = True

    ###### ページサイズ取得 //// ######

    ############################### PDFのページ数を取得 //// ##################################

    ######################### //// ドラッグアンドドロップ #########################
    def dragEnterEvent(self, e) -> None:
        if e.mimeData().hasUrls():
            e.accept()
        else:
            e.ignore()
        return super().dragEnterEvent(e)

    def dropEvent(self, e):
        tab_index = self.tabWidget.currentIndex()
        # tabWidgetの名前を取得
        tab_name = self.tabWidget.tabText(tab_index)
        target_count = len(e.mimeData().urls())
        first_target = e.mimeData().urls()[0].toLocalFile()
        is_pdf = False
        match tab_name:
            case 'PDF情報取得':
                if first_target.endswith('.pdf'):
                    list_or_label = self.page_count_listWidget
                    is_pdf = True
                if os.path.isdir(first_target):
                    if target_count == 1:
                        list_or_label = self.page_count_label
                    else:
                        QMessageBox.warning(self, 'Warning', 'フォルダを1つだけ選択してください')
                        return

        tsv_list = []
        if is_pdf:
            for url in e.mimeData().urls():
                file_path = str(url.toLocalFile())
                # file_pathの拡張子を取得
                ext = Path(file_path).suffix
                file_name = Path(file_path).name
                if ext == '.pdf':
                    # 複数ファイル可
                    if tab_name == 'PDF情報取得':
                        buf_list = []
                        for i in range(list_or_label.count()):
                            buf_text = list_or_label.item(i).text()
                            buf_text = buf_text.split('\t')[2]
                            buf_list.append(buf_text)
                        if tab_index == 0:
                            tsv_list.append('' + '\t' + file_name + '\t' + file_path)
                        else:
                            if not file_path in buf_list:
                                tsv_list.append('' + '\t' + file_name + '\t' + file_path)
        elif is_pdf == False:
            self.page_count_label.setText(first_target)

        if len(tsv_list) > 0 and is_pdf == True:
            list_or_label.addItems(tsv_list)
            self.set_index_to_listwidget(list_or_label) # myfunc

        super().dropEvent(e)

    # ファイルリストにインデックスを付ける
    def set_index_to_listwidget(self, listwidget_arg: QtWidgets.QListWidget):
        for i in range(listwidget_arg.count()):
            buf_item = listwidget_arg.item(i)
            buf_text = buf_item.text()
            a, b, c = buf_text.split('\t')
            a = str(i + 1)
            buf_text = '\t'.join([a, b, c])
            buf_item.setText(buf_text)

    ######################### ドラッグアンドドロップ //// #########################

########################## プログレスバークラス #########################
class ProgressBarDialog(QDialog, Ui_ProgressBarDialog):
    def __init__(self, max_value, *arg, **kwargs):
        super().__init__(*arg, **kwargs)
        self.setupUi(self)
        self.max_value = max_value
        self.progressBar: QtWidgets.QProgressBar
        self.last_update_time = time.time()
        flags = self.windowFlags()
        self.setWindowFlags(flags & ~Qt.WindowContextHelpButtonHint) # 除外
        self.setWindowFlags(flags | Qt.WindowStaysOnTopHint) # 追加

    # プログレスバーを更新する関数
    def update_progress_bar(self, current_value):
        current_time = time.time()
        if current_time - self.last_update_time > 0.2:
            percentage = (current_value / self.max_value) * 100
            self.progressBar.setValue(int(percentage))
            QCoreApplication.processEvents()
            self.last_update_time = current_time

####################### プログレスメッセージクラス #######################
class ProgressMsgDialog(QDialog, Ui_ProgressMsgDialog):
    def __init__(self, max_value, *arg, **kwargs):
        super().__init__(*arg, **kwargs)
        self.setupUi(self)
        self.max_value = max_value
        self.progress_msg_label: QtWidgets.QLabel
        flags = self.windowFlags()
        self.setWindowFlags(flags & ~Qt.WindowContextHelpButtonHint)
        self.setWindowFlags(flags | Qt.WindowStaysOnTopHint)

    def update_progress_msg(self, msg):
        # プログレスメッセージを更新する関数
        self.progress_msg_label.setText(msg)
        # プロセス更新
        QCoreApplication.processEvents()

############################ シングルトンクラス ############################
# シングルトンクラスを作成
class SingleApplication(QApplication):
    def __init__(self, argv):
        super().__init__(argv)
        self._memory = None

    def exec_(self):
        self._memory = QSharedMemory(self.applicationName())
        if self._memory.attach():
            # 他のインスタンスが存在する場合は終了
            return False
        self._memory.create(1)

        return super().exec_()

    def quit(self):
        if self._memory:
            self._memory.detach()
        super().quit()

####################### メイン処理 ###########################
if __name__ == '__main__':
    app = SingleApplication(sys.argv)
    # app = QApplication(sys.argv)
    icon_path = './data/icon_main_window.ico'
    if os.path.exists(icon_path):
        app.setWindowIcon(QtGui.QIcon(icon_path))
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())
