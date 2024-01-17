import os
import re
from pathlib import Path
import shutil
import sys
import time
from PyQt5 import QtGui, QtWidgets
from PyQt5.QtCore import Qt, QCoreApplication, QSharedMemory
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog, QDialog
import openpyxl
import datetime as dt
import json
from main_app_ui import Ui_MainWindow
import win32com.client as win32
from progress_bar_ui import Ui_ProgressBarDialog
from progress_msg_ui import Ui_ProgressMsgDialog

class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setAcceptDrops(True)
        self.init_config() # MyFunc

        self.show_listwidget(self.lineEdit_dir_path.text()) # MyFunc

        # インスタンス変数
        self.progress_cancel_flag = False
        # タイプヒント
        self.tabWidget: QtWidgets.QTabWidget
        self.listWidget_files_info: QtWidgets.QListWidget
        self.lineEdit_dir_path: QtWidgets.QLineEdit
        self.pushButton_label_clear: QtWidgets.QPushButton
        self.pushButton_list_clear: QtWidgets.QPushButton
        self.pushButton_dir_dialog: QtWidgets.QPushButton
        self.pushButton_exe_1: QtWidgets.QPushButton

        self.radioButton_copy: QtWidgets.QRadioButton
        self.radioButton_move: QtWidgets.QRadioButton
        self.lineEdit_rel_path: QtWidgets.QLineEdit
        self.pushButton_dir_dialog_2: QtWidgets.QPushButton

        self.pushButton_list_reload: QtWidgets.QPushButton
        self.pushButton_confirm: QtWidgets.QPushButton

        self.lineEdit_reqsec_show: QtWidgets.QLineEdit
        self.lineEdit_reqsec_input: QtWidgets.QLineEdit
        self.pushButton_reqsec: QtWidgets.QPushButton

        self.label_file_count: QtWidgets.QLabel

        self.lineEdit_file_exixts: QtWidgets.QLineEdit
        self.pushButton_file_exixts_exe: QtWidgets.QPushButton
        self.pushButton_file_exixts_dialog: QtWidgets.QPushButton

        # tabWidgetのインデックス0を選択
        self.tabWidget.setCurrentIndex(0)

        if self.lineEdit_reqsec_show.text() == '':
            self.lineEdit_reqsec_show.setText('0.25')

        ########################### イベントドリブン #############################

        # self.pushButton_exe_1.clicked.connect(lambda: self.temp_test())
        self.pushButton_exe_1.clicked.connect(lambda: self.main_exe())

        self.pushButton_confirm.clicked.connect(lambda: self.main_exe(is_confirm=True))

        self.pushButton_label_clear.clicked.connect(lambda: self.label_clear(self.lineEdit_dir_path))
        self.pushButton_list_clear.clicked.connect(lambda: self.listwidget_clear(self.listWidget_files_info))

        self.pushButton_dir_dialog.clicked.connect(lambda: self.set_dir_path(self.lineEdit_dir_path, True))
        self.pushButton_dir_dialog_2.clicked.connect(lambda: self.set_dir_path(self.lineEdit_rel_path))

        self.pushButton_list_reload.clicked.connect(lambda: self.show_listwidget(self.lineEdit_dir_path.text(), is_confirm=True))

        self.radioButton_copy.clicked.connect(lambda: self.auto_save_config_copy_or_move())
        self.radioButton_move.clicked.connect(lambda: self.auto_save_config_copy_or_move())

        self.pushButton_reqsec.clicked.connect(self.set_reqsec)

        self.pushButton_file_exixts_dialog.clicked.connect(lambda: self.set_file_path(self.lineEdit_file_exixts))
        self.pushButton_file_exixts_exe.clicked.connect(lambda: self.check_file_exists())

    ############################## //// Sub Process ############################

    # テスト用
    def temp_test(self):
        max = 30
        self.progress_bar_instance = ProgressBarDialog(self, max)
        self.progress_bar_instance.show()
        for i in range(max):
            self.progress_bar_instance.update_progress_bar(i)
            if self.progress_cancel_flag == True:
                QMessageBox.information(self, 'Info', 'キャンセルボタンが押されました')
                return
        self.progress_bar_instance.close()

    # 待機時間を設定
    def set_reqsec(self):
        if QMessageBox.question(self, 'Confirm', '待機時間を設定しますか？', QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return

        try:
            reqsec = float(self.lineEdit_reqsec_input.text())
        except:
            QMessageBox.warning(self, 'Warning', '数値を入力してください')
            return

        if reqsec < 0:
            QMessageBox.warning(self, 'Warning', '0以上の数値を入力してください')
            return
        reqsec = round(reqsec, 2)
        config = self.get_config()
        config['wait_sec'] = reqsec
        self.set_config(config)
        self.lineEdit_reqsec_show.setText(str(reqsec))
        QMessageBox.information(self, 'Info', '待機時間を設定しました')
        self.lineEdit_reqsec_input.setText('')

    # 設定自動保存
    def auto_save_config_copy_or_move(self):
        config = self.get_config()
        if self.radioButton_copy.isChecked():
            config['is_copy'] = True
        else:
            config['is_copy'] = False
        self.set_config(config)

    # ラベルをクリア
    def label_clear(self, target_label: QtWidgets.QLabel) -> None:
        if QMessageBox.question(self, 'Confirm', 'フォルダパスをクリアしますか？', QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return
        target_label.setText('')

    # listをすべてクリア
    def listwidget_clear(self, target_list: QtWidgets.QListWidget) -> None:
        if QMessageBox.question(self, 'Confirm', 'リストをクリアしますか？', QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return
        target_list.clear()

    # listの選択項目を削除
    def delete_one_item_from_listwidget(self):
        for item in self.listWidget_files_info.selectedItems():
            self.listWidget_files_info.takeItem(self.listWidget_files_info.row(item))
            if self.listWidget_files_info.count() != 0:
                self.listWidget_files_info.setCurrentRow(0)

    # プログレスバーを表示
    def show_progress_bar(self, max_value):
        self.progress_bar_instance = ProgressBarDialog(self, max_value)
        self.progress_bar_instance.show()
        QCoreApplication.processEvents()

    # 現在時刻文字列を取得する
    def get_datetime_str(self):
        return dt.datetime.now().strftime('%Y%m%d_%H%M_%S')

    # jsonファイルを読み込む
    def get_config(self):
        config_path = './data/config.json'
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config

    # jsonファイルに保存
    def set_config(self, config):
        config_path = './data/config.json'
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=4, ensure_ascii=False)

    # フォルダパスを設定する
    def set_dir_path(self, target_linewidget: QtWidgets.QLineEdit, is_show_files: bool = False) -> None:
        temp_path = QFileDialog.getExistingDirectory(self, 'フォルダを選択', './')
        if temp_path == '':
            return
        temp_path = os.path.normpath(temp_path)
        target_linewidget.setText(temp_path)
        # configに保存
        config = self.get_config()
        config['path']['dir_path'] = temp_path
        self.set_config(config)
        #
        if is_show_files == True:
            self.show_listwidget(temp_path)
        QMessageBox.information(self, 'Info', 'フォルダパスを設定しました')

    # ファイルパスを設定する
    def set_file_path(self, target_linewidget: QtWidgets.QLabel) -> None:
        if os.path.exists('./output') == False:
            init_path = os.getcwd()
        else:
            init_path = './output'
        temp_path, _ = QFileDialog.getOpenFileName(self, 'ファイルを選択', init_path, 'Excel Files (*.xlsx)')
        if temp_path == '':
            return
        target_linewidget.setText(temp_path)

    # ファイルリストを表示
    def show_listwidget(self, dir_path: str, is_confirm: bool = False):
        if is_confirm == True:
            if QMessageBox.question(self, 'Confirm', 'ファイルリストを更新しますか？', QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
                return
        if os.path.exists(dir_path) == False:
            error_msg_list = ['フォルダが存在しません > ' + dir_path]
            self.listWidget_files_info.addItems(error_msg_list)
            return
        self.listWidget_files_info.clear()
        temp_p = Path(dir_path)
        temp_list = []
        temp_p_list = list(temp_p.iterdir())
        for p in temp_p_list:
            # if p.is_dir():
            #     continue
            file_name = p.name
            temp_list.append(f'{file_name}')
        self.listWidget_files_info.addItems(temp_list)
        file_count = self.listWidget_files_info.count()
        self.label_file_count.setText(f'Count: {file_count}')
        if is_confirm == True:
            QMessageBox.information(self, 'Info', 'ファイルリストを更新しました')

    # 初期設定
    def init_config(self):
        config = self.get_config()
        self.lineEdit_dir_path.setText(config['path']['dir_path'])
        if config['is_copy'] == True:
            self.radioButton_copy.setChecked(True)
        else:
            self.radioButton_move.setChecked(True)
        self.lineEdit_rel_path.setText(config['path']['rel_path'])
        self.lineEdit_reqsec_show.setText(str(config['wait_sec']))


    # 設定を保存
    def save_config(self):
        if QMessageBox.question(self, 'Confirm', '設定を保存しますか？', QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return
        config = self.get_config()
        # フォルダパスを保存
        config['path']['dir_path'] = self.lineEdit_dir_path.text()
        config['path']['rel_path'] = self.lineEdit_rel_path.text()
        # ラジオボタン
        if self.radioButton_copy.isChecked():
            config['is_copy'] = True
        else:
            config['is_copy'] = False
        self.set_config(config)
        QMessageBox.information(self, 'Info', '設定を保存しました')

    # 終了処理

    def closeEvent(self, event):
        config = self.get_config()
        config['path']['rel_path'] = self.lineEdit_rel_path.text()
        self.set_config(config)
        event.accept()

    ############################## Sub Process //// ############################

    #######################################################################################
    ############################### //// Main Process ##################################
    #######################################################################################
    def main_exe(self, is_confirm: bool = False):
        if self.radioButton_copy.isChecked():
            is_copy = True
            mode = 'コピー'
        else:
            is_copy = False
            mode = '移動'
        mypath = self.lineEdit_dir_path.text()
        dest_root = self.lineEdit_rel_path.text()
        main_p = Path(mypath)
        p_list = list(main_p.iterdir())
        p_count = len(p_list)
        if is_confirm == False:
            if QMessageBox.question(self, 'Confirm', 'ファイル処理を実行しますか？\n\n'
                                    f'処理モード: {mode}\n\n'
                                    f'ターゲットパス:\n{mypath}\n\n'
                                    f'保存先ルートパス:\n{dest_root}\n\n'
                                    , QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
                return
        else:
            if QMessageBox.question(self, 'Confirm', '処理内容を確認しますか？\n\n'
                                    f'処理モード: {mode}\n\n'
                                    f'ターゲットパス:\n{mypath}\n\n'
                                    f'保存先ルートパス:\n{dest_root}\n\n'
                                    '※処理は実行されません\n'
                                    '※パーミッションエラーは検知できません'
                                    , QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
                return
        if os.path.exists(mypath) == False:
            QMessageBox.warning(self, 'Warning', 'コピー元フォルダが存在しません\n\n' + mypath)
            return
        start = time.perf_counter() # 処理時間計測開始

        if len(p_list) == 0:
            QMessageBox.warning(self, 'Warning', 'フォルダが空です')
            return

        # プログレスバーを表示
        self.show_progress_bar(p_count)

        progress_count = 0
        columns_name = ['file_name', 'message', 'folder_1', 'folder_2', 'folder_3', 'target_file_path', 'dest_file_path']
        result_list = []

        # 待機時間を取得
        reqsec = self.lineEdit_reqsec_show.text()
        try:
            reqsec = float(reqsec)
            reqsec = round(reqsec, 2)
        except:
            reqsec = 0.25

        for p in p_list:


            if p_count >=10000:
                if progress_count % 100 == 0:
                    self.progress_bar_instance.update_progress_bar(progress_count)
            elif p_count >= 1000:
                if progress_count % 10 == 0:
                    self.progress_bar_instance.update_progress_bar(progress_count)
            else:
                self.progress_bar_instance.update_progress_bar(progress_count)
            progress_count += 1

            file_path = str(p.absolute())
            file_name = p.name
            ext = p.suffix
            p1, p2, p3 = '', '', ''
            message = ''
            dest_path = ''
            dest_file_path = ''

            is_valid = True
            if file_name.count('_') < 4:
                message = 'アンダースコアの数が4未満'
                is_valid = False

            if is_valid: # パスが有効な場合
                p1 = file_name.split('_')[1]
                p2 = file_name.split('_')[2]
                p3 = file_name.split('_')[3]
                if p1 != '' and p2 != '' and p3 != '':
                    dest_path = os.path.join(dest_root, p1, p2, p3)
                elif p1 != '' and p2 != '' and p3 == '':
                    dest_path = os.path.join(dest_root, p1, p2)
                elif p1 != '' and p2 == '' and p3 == '':
                    dest_path = os.path.join(dest_root, p1)

                if p.is_file() == False:
                    message = 'ファイルではありません'
                    is_valid = False
                elif ext != '.pdf':
                    message = 'PDFファイルではありません'
                    is_valid = False
                elif os.path.exists(dest_path) == False:
                    message = '保存先フォルダが存在しません'
                    is_valid = False
                elif os.path.exists(os.path.join(dest_path, file_name)):
                    message = 'ファイルがすでに存在してます'
                    is_valid = False

            if is_valid == True: # パスが有効な場合
                if is_copy == True: # コピーの場合
                    try:
                        if is_confirm == False:
                            shutil.copy2(file_path, dest_path)
                            message = 'コピーしました'
                        else:
                            message = 'コピー可能'
                        dest_file_path = os.path.join(dest_path, file_name)
                        result_list.append([file_name, message, p1, p2, p3, file_path, dest_file_path])
                    except Exception as e: # ファイルにアクセスできないなどのエラー
                        message = e.strerror
                        p1, p2, p3 = '', '', ''
                        result_list.append([file_name, message, p1, p2, p3, file_path, dest_file_path])
                elif is_copy == False: # 移動の場合
                    try:
                        if is_confirm == False:
                            shutil.move(file_path, dest_path)
                            message = '移動しました'
                        else:
                            message = '移動可能'
                        dest_file_path = os.path.join(dest_path, file_name)
                        result_list.append([file_name, message, p1, p2, p3, file_path, dest_file_path])
                    except Exception as e: # ファイルにアクセスできないなどのエラー
                        message = e.strerror
                        p1, p2, p3 = '', '', ''
                        result_list.append([file_name, message, p1, p2, p3, file_path, dest_file_path])
            elif is_valid == False: # パスが無効な場合
                p1, p2, p3, dest_file_path = '', '', '', ''
                result_list.append([file_name, message, p1, p2, p3, file_path, dest_file_path])

            # コピー／移動モードの場合、かつ、処理が有効だった場合 待機する
            if is_confirm == False and is_valid == True:
                time.sleep(reqsec)

        # プログレスバーを閉じる
        self.progress_bar_instance.close()

        end = time.perf_counter()
        sec = round(end - start, 2)
        if is_confirm == False:
            QMessageBox.information(self, 'Info','処理が完了しました\n\n' f'処理時間: {sec}秒')
        # 結果をエクセルに書き出し
        self.write_excel(result_list, columns_name, is_confirm)

    # エクセルに書き出し
    def write_excel(self, result_list: list, columns_name: list, is_confirm: bool) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active

        for i in range(len(columns_name)):
            ws.cell(row=1, column=i + 1).value = columns_name[i]
        for row in range(len(result_list)):
            for col in range(len(result_list[0])):
                ws.cell(row=row + 2, column=col + 1).value = result_list[row][col]
        dt_str = dt.datetime.now().strftime('%Y%m%d_%H%M_%S')
        excel_path = f'./output/_log_{dt_str}.xlsx'

        # 列の自動調整
        self.auto_fit_column(ws) # MyFunc

        excel_full_path = os.path.join(os.getcwd(), excel_path)
        ###
        # ファイル存在確認用ラインに設定
        if is_confirm == False:
            self.lineEdit_file_exixts.setText(excel_full_path)
        ###

        # outputフォルダがなければ作成
        if os.path.exists('./output') == False:
            os.makedirs('./output')
        if os.path.exists(excel_full_path) == False:
            wb.save(excel_full_path)
            if QMessageBox.question(self, 'Confirm',
                                    f'{excel_full_path}\nを作成しました\n\n'
                                    'エクセルファイルを開きますか？',
                                    QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                # 保存したエクセルファイルを開く
                # pro = Popen(['start', 'excel', excel_full_path], shell=True)
                self.open_excel_file(excel_full_path)

                ## os.startfile(excel_full_path)
            else:
                return
        else:
            QMessageBox.warning(self, 'Warning', 'Exctlファイルを書き出せませんでした')
            return

    def open_excel_file(self, excel_file_path: str) -> None:
        excel_app = win32.gencache.EnsureDispatch('Excel.Application')
        excel_app.Workbooks.Open(excel_file_path)
        excel_app.Visible = True

    # 列の自動調整
    def auto_fit_column(self, ws):
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjust_width = int((max_length + 5) * 1.1)
            ws.column_dimensions[column_letter].width = adjust_width

    # ログファイルでファイルの存在確認を行う
    def check_file_exists(self):
        if QMessageBox.question(self, 'Confirm', 'ファイルの存在確認を行いますか？', QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return

        temp_path = self.lineEdit_file_exixts.text()
        if os.path.exists(temp_path) == False:
            QMessageBox.warning(self, 'Warning', 'ファイルが存在しません')
            return
        wb = openpyxl.load_workbook(temp_path)
        ws = wb.active
        if ws.cell(row=1, column=7).value != 'dest_file_path':
            QMessageBox.warning(self, 'Warning', 'セルG1に文字列 dest_file_path がありません')
            return
        ws.cell(row=1, column=8, value='isExists')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
            for cell in row:
                ws.cell(row=cell.row, column=8, value='')
                if cell.value is not None:
                    if os.path.exists(cell.value):
                        ws.cell(row=cell.row, column=8, value=True)
        parent_path = os.path.dirname(temp_path)

        name_part = Path(temp_path).stem
        exi_part = Path(temp_path).suffix
        next_name = name_part + '_checked' + exi_part
        next_path = os.path.join(parent_path, next_name)
        wb.save(next_path)

        if QMessageBox.question(self, 'Confirm', 'ファイルを開きますか？', QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            # 保存したエクセルファイルを開く
            self.open_excel_file(next_path)

    ###################################################################################
    ############################### Main Process //// ##################################
    ###################################################################################

    ######################### //// ドラッグアンドドロップ #########################
    def dragEnterEvent(self, e) -> None:
        if e.mimeData().hasUrls():
            e.accept()
        else:
            e.ignore()
        return super().dragEnterEvent(e)

    def dropEvent(self, e):
        tab_index = self.tabWidget.currentIndex()
        # tabWidgetの名前を取得
        tab_name = self.tabWidget.tabText(tab_index)
        target_count = len(e.mimeData().urls())
        first_target = e.mimeData().urls()[0].toLocalFile()
        match tab_name:
            case 'メイン':
                if os.path.isdir(first_target):
                    if target_count == 1:
                        self.lineEdit_dir_path.setText(first_target)
                        self.show_listwidget(first_target)
                        config = self.get_config()
                        config['path']['dir_path'] = first_target
                        self.set_config(config)
                    else:
                        QMessageBox.warning(self, 'Warning', 'フォルダを1つだけ選択してください')
                        return
            case '設定':
                if os.path.isdir(first_target):
                    if target_count == 1:
                        self.lineEdit_rel_path.setText(first_target)
                        config = self.get_config()
                        config['path']['rel_path'] = first_target
                        self.set_config(config)
                    else:
                        QMessageBox.warning(self, 'Warning', 'フォルダを1つだけ選択してください')
                        return
            case '存在確認':
                if os.path.isfile(first_target):
                    if target_count == 1:
                        temp_p = Path(first_target)
                        ext = temp_p.suffix
                        if ext == '.xlsx':
                            self.lineEdit_file_exixts.setText(first_target)
                        else:
                            QMessageBox.warning(self, 'Warning', 'エクセルファイルを選択してください')
                            return
                    else:
                        QMessageBox.warning(self, 'Warning', 'エクセルファイルを1つだけ選択してください')
                        return
                else:
                    QMessageBox.warning(self, 'Warning', 'エクセルファイルを選択してください')
                    return
        super().dropEvent(e)

    ######################### ドラッグアンドドロップ //// #########################

########################## プログレスバークラス #########################
class ProgressBarDialog(QDialog, Ui_ProgressBarDialog):
    def __init__(self, main_window,  max_value, *arg, **kwargs):
        super().__init__(*arg, **kwargs)
        self.setupUi(self)
        self.main_window = main_window
        self.max_value = max_value

        self.main_window.progress_cancel_flag = False

        self.progressBar: QtWidgets.QProgressBar
        self.label: QtWidgets.QLabel
        self.pushButton_cancel: QtWidgets.QPushButton

        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint) # 除外
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint) # 追加

        self.pushButton_cancel.clicked.connect(lambda: self.cancel())

    # プログレスバーを更新する関数
    def update_progress_bar(self, current_value):
        percentage = (current_value / self.max_value) * 100
        self.progressBar.setValue(int(percentage))
        QCoreApplication.processEvents()

    # キャンセルボタンを押したときの処理
    def cancel(self):
        self.main_window.progress_cancel_flag = True
        self.close()

####################### プログレスメッセージクラス #######################
class ProgressMsgDialog(QDialog, Ui_ProgressMsgDialog):
    def __init__(self, max_value, *arg, **kwargs):
        super().__init__(*arg, **kwargs)
        self.setupUi(self)
        self.max_value = max_value
        self.progress_msg_label: QtWidgets.QLabel

        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint) # 除外
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint) # 追加

    def update_progress_msg(self, msg):
        # プログレスメッセージを更新する関数
        self.progress_msg_label.setText(msg)
        # プロセス更新
        QCoreApplication.processEvents()
############################ シングルトンクラス ############################
# シングルトンクラスを作成
class SingleApplication(QApplication):
    def __init__(self, argv):
        super().__init__(argv)
        self._memory = None

    def exec_(self):
        self._memory = QSharedMemory(self.applicationName())
        if self._memory.attach():
            # 他のインスタンスが存在する場合は終了
            return False
        self._memory.create(1)

        return super().exec_()

    def quit(self):
        if self._memory:
            self._memory.detach()
        super().quit()

####################### メイン処理 ###########################
if __name__ == '__main__':
    app = SingleApplication(sys.argv)
    # app = QApplication(sys.argv)
    icon_path = './data/icon_main_window.ico'
    if os.path.exists(icon_path):
        app.setWindowIcon(QtGui.QIcon(icon_path))
    main_window = MainWindow()
    if True:
        main_window.show()
        sys.exit(app.exec_())
